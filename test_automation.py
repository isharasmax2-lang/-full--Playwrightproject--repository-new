from playwright.sync_api import sync_playwright
import time
import os
import argparse
import re
from pathlib import Path
import sys
import openpyxl
from openpyxl.cell.cell import MergedCell
from html import escape

# Configuration
ROOT_DIR = Path(__file__).resolve().parent.parent
TESTS_DIR = ROOT_DIR / "test_automation"

DEFAULT_EXCEL_CANDIDATES = [
    str(TESTS_DIR / "Assignment 1 - Test cases.xlsx"),
]

DEFAULT_SHEET_NAME = " Test cases"
DEFAULT_FRONTEND_URL = os.getenv("FRONTEND_URL", "https://www.pixelssuite.com/chat-translator")

DEFAULT_INPUT_COLUMN_CANDIDATES = [
    "Singlish",
    "Input",
    "Singlish Input",
    "Test Input",
    "Source",
    "Sentence",
    "Text",
]

DEFAULT_EXPECTED_COLUMN_CANDIDATES = [
    "Sinhala",
    "Expected_Output",
    "Expected Output",
    "Expected output",
    "Expected",
    "Expected Sinhala",
]

DEFAULT_ACTUAL_COLUMN_CANDIDATES = [
    "Actual_Output",
    "Actual Output",
    "Actual output",
    "Actual",
]

DEFAULT_STATUS_COLUMN_CANDIDATES = [
    "Status",
    "Result",
    "Pass/Fail",
    "Pass Fail",
]

DEFAULT_WAIT_MS = 5000
DEFAULT_RETRIES = 8
DEFAULT_RETRY_WAIT_MS = 1000
DEFAULT_TYPE_DELAY_MS = 30
DEFAULT_TIMEOUT_MS = 60000
DEFAULT_SLOW_MO_MS = 0
DEFAULT_RESULTS_DIR = TESTS_DIR / "results"

def _safe_slug(value: str, fallback: str = "case") -> str:
    text = re.sub(r"[^a-zA-Z0-9]+", "-", str(value).strip().lower()).strip("-")
    return text or fallback

def _build_report_html(title: str, summary: dict, rows: list[dict]) -> str:
    status_counts = summary.get("status_counts", {})
    total = summary.get("total", 0)
    passed = summary.get("passed", 0)
    failed = summary.get("failed", 0)
    collected = summary.get("collected", 0)
    ui_errors = summary.get("ui_errors", 0)
    screenshots_dir = summary.get("screenshots_dir", "")

    def _card_class(status: str) -> str:
        return status.strip().lower().replace(" ", "-")

    cards = "".join(
        f"""
        <div class='card {_card_class(row['status'])}'>
            <div class='card-head'>
                <span class='badge'>{escape(row['status'])}</span>
                <span class='row-id'>Row {row['row_index']}</span>
            </div>
            <div class='case-title'>{escape(row['title'])}</div>
            <div class='kv'><span>Expected</span><div>{escape(row['expected'] or '-')}</div></div>
            <div class='kv'><span>Actual</span><div>{escape(row['actual'] or '-')}</div></div>
            <div class='kv'><span>Screenshot</span><div><a href='{escape(row['screenshot_rel'])}' target='_blank'>{escape(row['screenshot_name'])}</a></div></div>
            <a class='thumb-link' href='{escape(row['screenshot_rel'])}' target='_blank'>
                <img src='{escape(row['screenshot_rel'])}' alt='screenshot for row {row['row_index']}' />
            </a>
        </div>
        """
        for row in rows
    )

    return f"""<!doctype html>
<html lang='en'>
<head>
  <meta charset='utf-8' />
  <meta name='viewport' content='width=device-width, initial-scale=1' />
  <title>{escape(title)}</title>
  <style>
    :root {{
      color-scheme: dark;
      --bg: #0b1020;
      --panel: #11182d;
      --panel-2: #16213c;
      --text: #e5eefc;
      --muted: #9db0d0;
      --pass: #22c55e;
      --fail: #ef4444;
      --collect: #eab308;
      --error: #f97316;
      --border: rgba(148, 163, 184, 0.2);
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Segoe UI, Arial, sans-serif; background: radial-gradient(circle at top, #152241 0%, var(--bg) 58%); color: var(--text); }}
    .wrap {{ max-width: 1400px; margin: 0 auto; padding: 28px 20px 40px; }}
    .hero {{ background: linear-gradient(135deg, rgba(34,197,94,.18), rgba(59,130,246,.16)); border: 1px solid var(--border); border-radius: 20px; padding: 22px; box-shadow: 0 18px 40px rgba(0,0,0,.25); }}
    h1 {{ margin: 0 0 10px; font-size: 28px; }}
    .meta {{ color: var(--muted); display: flex; flex-wrap: wrap; gap: 14px; font-size: 14px; }}
    .stats {{ display: grid; grid-template-columns: repeat(5, minmax(0, 1fr)); gap: 12px; margin-top: 16px; }}
    .stat {{ background: rgba(15, 23, 42, .72); border: 1px solid var(--border); border-radius: 16px; padding: 16px; }}
    .stat .label {{ color: var(--muted); font-size: 12px; text-transform: uppercase; letter-spacing: .08em; }}
    .stat .value {{ font-size: 26px; font-weight: 700; margin-top: 6px; }}
    .section {{ margin-top: 24px; }}
    .section h2 {{ margin: 0 0 12px; font-size: 18px; }}
    .grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(350px, 1fr)); gap: 16px; }}
    .card {{ background: rgba(15, 23, 42, .82); border: 1px solid var(--border); border-radius: 18px; padding: 16px; overflow: hidden; }}
    .card.pass {{ box-shadow: inset 0 0 0 1px rgba(34, 197, 94, .18); }}
    .card.fail {{ box-shadow: inset 0 0 0 1px rgba(239, 68, 68, .18); }}
    .card.collected {{ box-shadow: inset 0 0 0 1px rgba(234, 179, 8, .18); }}
    .card.ui-error {{ box-shadow: inset 0 0 0 1px rgba(249, 115, 22, .18); }}
    .card-head {{ display: flex; justify-content: space-between; gap: 10px; align-items: center; margin-bottom: 10px; }}
    .badge {{ display: inline-flex; align-items: center; padding: 4px 10px; border-radius: 999px; font-weight: 700; font-size: 12px; }}
    .pass .badge {{ background: rgba(34, 197, 94, .16); color: #86efac; }}
    .fail .badge {{ background: rgba(239, 68, 68, .16); color: #fca5a5; }}
    .collected .badge {{ background: rgba(234, 179, 8, .16); color: #fde68a; }}
    .ui-error .badge {{ background: rgba(249, 115, 22, .16); color: #fdba74; }}
    .row-id {{ color: var(--muted); font-size: 13px; }}
    .case-title {{ font-weight: 700; line-height: 1.4; margin-bottom: 12px; }}
    .kv {{ display: grid; grid-template-columns: 92px 1fr; gap: 12px; margin: 8px 0; font-size: 13px; }}
    .kv span {{ color: var(--muted); text-transform: uppercase; letter-spacing: .05em; font-size: 11px; }}
    .thumb-link {{ display: block; margin-top: 12px; }}
    .thumb-link img {{ width: 100%; border-radius: 12px; border: 1px solid var(--border); background: #0f172a; }}
    a {{ color: #93c5fd; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    .footer {{ margin-top: 18px; color: var(--muted); font-size: 13px; }}
    @media (max-width: 900px) {{ .stats {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }} .grid {{ grid-template-columns: 1fr; }} }}
  </style>
</head>
<body>
  <div class='wrap'>
    <div class='hero'>
      <h1>{escape(title)}</h1>
      <div class='meta'>
        <span>Total cases: {total}</span>
        <span>Passed: {passed}</span>
        <span>Failed: {failed}</span>
        <span>Collected: {collected}</span>
        <span>UI errors: {ui_errors}</span>
      </div>
      <div class='stats'>
        <div class='stat'><div class='label'>Total</div><div class='value'>{total}</div></div>
        <div class='stat'><div class='label'>Passed</div><div class='value' style='color: var(--pass);'>{passed}</div></div>
        <div class='stat'><div class='label'>Failed</div><div class='value' style='color: var(--fail);'>{failed}</div></div>
        <div class='stat'><div class='label'>Collected</div><div class='value' style='color: var(--collect);'>{collected}</div></div>
        <div class='stat'><div class='label'>UI Errors</div><div class='value' style='color: var(--error);'>{ui_errors}</div></div>
      </div>
    </div>

    <div class='section'>
      <h2>Test Case Evidence</h2>
      <div class='grid'>
        {cards}
      </div>
      <div class='footer'>Screenshots directory: {escape(screenshots_dir)}</div>
      <div class='footer'>Status counts: {escape(str(status_counts))}</div>
    </div>
  </div>
</body>
</html>"""

def _configure_stdout():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass

def _pick_existing_path(candidates):
    for p in candidates:
        if p and os.path.exists(p):
            return p
    return candidates[0] if candidates else None

def _resolve_path(p: str | None) -> str | None:
    if not p:
        return None
    path = Path(p)
    if path.is_absolute():
        if path.exists():
            return str(path)

        normalized_target = re.sub(r"\s+", " ", path.name).strip().lower()
        search_dirs = [path.parent]
        for directory in search_dirs:
            if not directory.exists():
                continue
            for candidate in directory.iterdir():
                if not candidate.is_file():
                    continue
                candidate_name = re.sub(r"\s+", " ", candidate.name).strip().lower()
                if candidate_name == normalized_target:
                    return str(candidate)
                if candidate.suffix.lower() == path.suffix.lower() and re.sub(r"\s+", "", candidate.stem).lower() == re.sub(r"\s+", "", path.stem).lower():
                    return str(candidate)

        return str(path)
    root_candidate = (ROOT_DIR / path).resolve()
    if root_candidate.exists():
        return str(root_candidate)
    tests_candidate = (TESTS_DIR / path).resolve()
    if tests_candidate.exists():
        return str(tests_candidate)

    normalized_target = re.sub(r"\s+", " ", path.name).strip().lower()
    for directory in {ROOT_DIR, TESTS_DIR, path.parent if path.parent != Path("") else None}:
        if not directory or not directory.exists():
            continue
        for candidate in directory.iterdir():
            if not candidate.is_file():
                continue
            candidate_name = re.sub(r"\s+", " ", candidate.name).strip().lower()
            if candidate_name == normalized_target:
                return str(candidate)
            if candidate.suffix.lower() == path.suffix.lower() and re.sub(r"\s+", "", candidate.stem).lower() == re.sub(r"\s+", "", path.stem).lower():
                return str(candidate)

    return str(root_candidate)

def _normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())

def _header_values(ws, row_index: int) -> list:
    max_col = max(1, int(ws.max_column or 1))
    return [ws.cell(row=row_index, column=c).value for c in range(1, max_col + 1)]

def _find_header_row(ws, max_scan_rows: int) -> int:
    input_tokens = {_normalize_header(v) for v in DEFAULT_INPUT_COLUMN_CANDIDATES}
    expected_tokens = {_normalize_header(v) for v in DEFAULT_EXPECTED_COLUMN_CANDIDATES}
    actual_tokens = {_normalize_header(v) for v in DEFAULT_ACTUAL_COLUMN_CANDIDATES}
    status_tokens = {_normalize_header(v) for v in DEFAULT_STATUS_COLUMN_CANDIDATES}

    best_score = -1
    best_row = 1
    scan_limit = max(1, min(int(max_scan_rows), int(ws.max_row or 1)))
    for r in range(1, scan_limit + 1):
        values = _header_values(ws, r)
        texts = [v for v in values if isinstance(v, str) and v.strip() and len(v.strip()) <= 40]
        if len(texts) < 2:
            continue

        norms = {_normalize_header(v) for v in texts}
        if "tcid" in norms and "input" in norms and "expectedoutput" in norms:
            return r

        if "input" not in norms:
            continue
        if not (norms & expected_tokens):
            continue

        score = 0
        for v in texts:
            n = _normalize_header(v)
            if n in input_tokens:
                score += 3
            if n in expected_tokens:
                score += 2
            if n in actual_tokens:
                score += 1
            if n in status_tokens:
                score += 1
        if score > best_score:
            best_score = score
            best_row = r
    return best_row

def _merged_top_left_cell(ws, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return ws.cell(row=row, column=col)

def _is_top_left_of_merged_cell(ws, row: int, col: int) -> bool:
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return True
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row == row and rng.min_col == col
    return True

def _set_cell_value(ws, row: int, col: int, value):
    cell = _merged_top_left_cell(ws, row, col)
    cell.value = value

def _find_column_index(header_values: list, requested_name: str | None, candidates: list[str]) -> int | None:
    indexed = []
    for i, v in enumerate(header_values, start=1):
        if v is None:
            continue
        indexed.append((i, str(v)))

    norm_to_index: dict[str, int] = {}
    for i, v in indexed:
        n = _normalize_header(v)
        if n and n not in norm_to_index:
            norm_to_index[n] = i

    def match(name: str) -> int | None:
        n = _normalize_header(name)
        if not n:
            return None
        if n in norm_to_index:
            return norm_to_index[n]
        for i, v in indexed:
            if n in _normalize_header(v) or _normalize_header(v) in n:
                return i
        return None

    if requested_name:
        found = match(requested_name)
        if found:
            return found

    for c in candidates:
        found = match(c)
        if found:
            return found

    return None

def _last_header_col(header_values: list) -> int:
    last = 0
    for i, v in enumerate(header_values, start=1):
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        last = i
    return last

def _ensure_column(ws, header_row: int, header_values: list, desired_name: str) -> int:
    found = _find_column_index(header_values, desired_name, [])
    if found:
        return found
    col = _last_header_col(header_values) + 1
    ws.cell(row=header_row, column=col).value = desired_name
    if col <= len(header_values):
        header_values[col - 1] = desired_name
    else:
        while len(header_values) < col - 1:
            header_values.append(None)
        header_values.append(desired_name)
    return col

def _dismiss_overlays(page):
    candidates = [
        ("button", re.compile(r"^(Accept|I Agree|Agree|OK|Got it)$", re.IGNORECASE)),
        ("button", re.compile(r"^(Accept all|Accept All)$", re.IGNORECASE)),
    ]
    for role, name in candidates:
        try:
            btn = page.get_by_role(role, name=name).first
            if btn.is_visible():
                btn.click(timeout=2000)
                page.wait_for_timeout(500)
        except Exception:
            pass

def _clear_textarea(page, locator, attempts: int = 3):
    for _ in range(max(1, int(attempts))):
        try:
            locator.focus(timeout=2000)
        except Exception:
            pass
        try:
            page.keyboard.press("Control+A")
            page.keyboard.press("Backspace")
        except Exception:
            pass
        try:
            locator.fill("")
        except Exception:
            pass
        try:
            locator.evaluate(
                """(el) => {
                    el.focus();
                    if ('value' in el) {
                        el.value = '';
                        el.dispatchEvent(new Event('input', { bubbles: true }));
                        el.dispatchEvent(new Event('change', { bubbles: true }));
                    }
                }"""
            )
        except Exception:
            pass
        try:
            if locator.input_value() == "":
                return
        except Exception:
            pass
        try:
            if locator.input_value() == "":
                return
        except Exception:
            pass
        page.wait_for_timeout(200)

def _ensure_input_value(page, input_locator, text: str, type_delay_ms: int):
    _clear_textarea(page, input_locator)
    if type_delay_ms and int(type_delay_ms) > 0:
        try:
            input_locator.focus(timeout=2000)
        except Exception:
            pass
        input_locator.type(text, delay=int(type_delay_ms))
    else:
        input_locator.fill(text)
    try:
        current = input_locator.input_value()
        if current is None:
            return
        if str(current).strip() == text.strip():
            return
    except Exception:
        return
    page.wait_for_timeout(150)
    _clear_textarea(page, input_locator)
    input_locator.fill(text)

def _read_output(is_chat: bool, output_locator) -> str:
    if is_chat:
        try:
            v = output_locator.input_value()
            if v is not None:
                v = str(v).strip()
                if v:
                    return v
        except Exception:
            pass
    try:
        v = output_locator.inner_text()
        if v is not None:
            v = str(v).strip()
            if v:
                return v
    except Exception:
        pass
    try:
        v = output_locator.text_content()
        if v is not None:
            v = str(v).strip()
            if v:
                return v
    except Exception:
        pass
    try:
        v = output_locator.evaluate("(el) => el && ('value' in el ? el.value : '')")
        if v is not None:
            v = str(v).strip()
            if v:
                return v
    except Exception:
        pass
    return ""

def _find_chat_locators(page, timeout_ms: int):
    deadline = time.time() + (max(1, timeout_ms) / 1000)
    last_debug = None
    while time.time() < deadline:
        _dismiss_overlays(page)
        try:
            input_by_ph = page.locator('textarea[placeholder*="English"]').first
            output_by_ph = page.locator('textarea[placeholder*="Sinhala"]').first
            if input_by_ph.count() > 0 and output_by_ph.count() > 0 and input_by_ph.is_visible() and output_by_ph.is_visible():
                action = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.IGNORECASE)).first
                return input_by_ph, output_by_ph, action
        except Exception:
            pass

        try:
            count = page.locator("textarea").count()
            visible = []
            for i in range(count):
                loc = page.locator("textarea").nth(i)
                if loc.is_visible():
                    visible.append(loc)
            if len(visible) >= 2:
                action = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.IGNORECASE)).first
                return visible[0], visible[1], action
        except Exception as e:
            last_debug = str(e)

        page.wait_for_timeout(500)

    try:
        meta = page.evaluate(
            """() => Array.from(document.querySelectorAll('textarea')).map(t => ({
              placeholder: t.getAttribute('placeholder') || '',
              disabled: !!t.disabled,
              readOnly: !!t.readOnly,
              visible: !!(t.offsetParent)
            }))"""
        )
        print("Debug: textarea meta:", meta)
    except Exception as e:
        print("Debug: failed to read textarea meta:", e)
    if last_debug:
        print("Debug: last error:", last_debug)
    raise RuntimeError("Could not find Chat UI locators (input/output textareas).")

def _parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=_pick_existing_path(DEFAULT_EXCEL_CANDIDATES))
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--header-row", type=int, default=0)
    parser.add_argument("--max-header-scan-rows", type=int, default=30)
    parser.add_argument("--input-col", default=None)
    parser.add_argument("--expected-col", default=None)
    parser.add_argument("--actual-col", default=None)
    parser.add_argument("--status-col", default=None)
    parser.add_argument("--url", default=DEFAULT_FRONTEND_URL)
    parser.add_argument("--output", default=None)
    parser.add_argument("--save-every", type=int, default=0)
    parser.add_argument("--headless", action="store_true", default=False)
    parser.add_argument("--wait-ms", type=int, default=DEFAULT_WAIT_MS)
    parser.add_argument("--retries", type=int, default=DEFAULT_RETRIES)
    parser.add_argument("--retry-wait-ms", type=int, default=DEFAULT_RETRY_WAIT_MS)
    parser.add_argument("--type-delay-ms", type=int, default=DEFAULT_TYPE_DELAY_MS)
    parser.add_argument("--timeout-ms", type=int, default=DEFAULT_TIMEOUT_MS)
    parser.add_argument("--slow-mo-ms", type=int, default=DEFAULT_SLOW_MO_MS)
    parser.add_argument("--results-dir", default=str(DEFAULT_RESULTS_DIR))
    parser.add_argument("--keep-open", action="store_true", default=False)
    return parser.parse_args()

def run_test():
    _configure_stdout()
    args = _parse_args()
    args.excel = _resolve_path(args.excel)
    args.output = _resolve_path(args.output) if args.output else args.excel
    results_dir = Path(_resolve_path(args.results_dir) or args.results_dir)
    screenshots_dir = results_dir / "screenshots"
    results_dir.mkdir(parents=True, exist_ok=True)
    screenshots_dir.mkdir(parents=True, exist_ok=True)

    if not args.excel or not os.path.exists(args.excel):
        print(f"Error: File '{args.excel}' not found.")
        return

    try:
        wb = openpyxl.load_workbook(args.excel)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    if args.sheet and args.sheet in wb.sheetnames:
        ws = wb[args.sheet]
    else:
        ws = wb.active

    header_row = int(args.header_row or 0)
    if header_row <= 0:
        header_row = _find_header_row(ws, int(args.max_header_scan_rows))

    header_values = _header_values(ws, header_row)

    input_col_idx = _find_column_index(header_values, args.input_col, DEFAULT_INPUT_COLUMN_CANDIDATES)
    expected_col_idx = _find_column_index(header_values, args.expected_col, DEFAULT_EXPECTED_COLUMN_CANDIDATES)

    if not input_col_idx:
        printable = [str(v) if v is not None else "" for v in header_values]
        print("Error: Could not resolve input column.")
        print(f"Header row: {header_row}")
        print(f"Available columns: {printable}")
        return

    actual_col_name = args.actual_col or "Actual output"
    status_col_name = args.status_col or "Status"

    actual_col_idx = _find_column_index(header_values, args.actual_col, DEFAULT_ACTUAL_COLUMN_CANDIDATES)
    status_col_idx = _find_column_index(header_values, args.status_col, DEFAULT_STATUS_COLUMN_CANDIDATES)

    actual_col_idx = actual_col_idx or _ensure_column(ws, header_row, header_values, actual_col_name)
    status_col_idx = status_col_idx or _ensure_column(ws, header_row, header_values, status_col_name)

    rows_total = max(0, int(ws.max_row or 0) - header_row)
    print(f"Starting Frontend-Only test with {rows_total} rows...")

    report_rows = []
    summary = {
        "total": 0,
        "passed": 0,
        "failed": 0,
        "collected": 0,
        "ui_errors": 0,
        "status_counts": {},
        "screenshots_dir": str(screenshots_dir),
    }

    with sync_playwright() as p:
        # 2. Launch Browser
        if args.headless:
            print("Running in headless mode: browser UI will not be visible. Remove --headless to watch typing.")
        browser = p.chromium.launch(headless=args.headless, slow_mo=max(0, int(args.slow_mo_ms)))
        page = browser.new_page()
        page.set_default_timeout(max(1000, int(args.timeout_ms)))

        # 3. Open Frontend
        try:
            page.goto(args.url, wait_until="domcontentloaded")
            try:
                page.wait_for_load_state("networkidle", timeout=max(1000, int(args.timeout_ms)))
            except Exception:
                pass
            page.wait_for_selector("textarea", timeout=max(1000, int(args.timeout_ms)))
            print("Frontend loaded successfully.")
        except Exception as e:
            print(f"Error loading frontend: {e}")
            browser.close()
            return

        is_chat = "chat-translator" in (args.url or "")
        if is_chat:
            try:
                input_locator, output_locator, action_locator = _find_chat_locators(page, int(args.timeout_ms))
            except Exception as e:
                print(f"Error locating chat UI elements: {e}")
                browser.close()
                return
        else:
            input_locator = page.locator("textarea")
            output_locator = page.locator("div.card").filter(has_text=re.compile(r"\\bSinhala\\b")).locator("div.bg-slate-50").first
            action_locator = None

        # 4. Iterate Rows
        processed = 0
        for row_index in range(header_row + 1, int(ws.max_row or 0) + 1):
            if not _is_top_left_of_merged_cell(ws, row_index, input_col_idx):
                continue

            input_cell = _merged_top_left_cell(ws, row_index, input_col_idx)
            input_value = input_cell.value
            singlish_input = str(input_value).strip() if input_value is not None else ""
            if not singlish_input:
                continue

            expected_value = (
                _merged_top_left_cell(ws, row_index, expected_col_idx).value if expected_col_idx else None
            )
            expected_sinhala = str(expected_value).strip() if expected_value is not None else ""

            print(f"Testing [Row {row_index}]: {singlish_input}")

            screenshot_name = f"row_{row_index}_{_safe_slug(singlish_input)}.png"
            screenshot_path = screenshots_dir / screenshot_name
            screenshot_rel = os.path.relpath(screenshot_path, results_dir).replace("\\", "/")
            current_status = "UI Error"
            actual_output = ""

            try:
                _dismiss_overlays(page)
                prev_output = _read_output(is_chat, output_locator)
                _ensure_input_value(page, input_locator, singlish_input, int(args.type_delay_ms))

                if action_locator:
                    action_locator.click()

                page.wait_for_timeout(max(0, int(args.wait_ms)))
                
                # Wait for visible content - retry a few times if empty
                tries = max(1, int(args.retries))
                for i in range(tries):
                    current = _read_output(is_chat, output_locator)
                    if not current:
                        page.wait_for_timeout(max(0, int(args.retry_wait_ms)))
                        continue
                    if prev_output and current == prev_output:
                        page.wait_for_timeout(max(0, int(args.retry_wait_ms)))
                        continue
                    if current:
                        actual_output = current
                        break
                    page.wait_for_timeout(max(0, int(args.retry_wait_ms)))

                if prev_output and actual_output == "" and prev_output != "":
                    raise RuntimeError("Output did not update for this input (still showing previous output).")

                _set_cell_value(ws, row_index, actual_col_idx, actual_output)

                if expected_sinhala:
                    current_status = "PASS" if actual_output == expected_sinhala else "FAIL"
                else:
                    current_status = "COLLECTED"
                _set_cell_value(ws, row_index, status_col_idx, current_status)
                print(f"  -> {current_status}")
                processed += 1
                if args.save_every and int(args.save_every) > 0 and processed % int(args.save_every) == 0:
                    wb.save(args.output)
                
            except Exception as e:
                print(f"Error in UI interaction: {e}")
                try:
                    _set_cell_value(ws, row_index, status_col_idx, "UI Error")
                except Exception:
                    pass
                if args.save_every and int(args.save_every) > 0:
                    try:
                        wb.save(args.output)
                    except Exception:
                        pass
                current_status = "UI Error"
                actual_output = f"Error: {e}"

            try:
                page.screenshot(path=str(screenshot_path), full_page=True)
            except Exception as screenshot_error:
                print(f"Warning: could not save screenshot for row {row_index}: {screenshot_error}")

            summary["total"] += 1
            status_key = current_status.lower().replace(" ", "_")
            summary["status_counts"][current_status] = summary["status_counts"].get(current_status, 0) + 1
            if current_status == "PASS":
                summary["passed"] += 1
            elif current_status == "FAIL":
                summary["failed"] += 1
            elif current_status == "COLLECTED":
                summary["collected"] += 1
            elif current_status == "UI Error":
                summary["ui_errors"] += 1

            report_rows.append(
                {
                    "row_index": row_index,
                    "title": singlish_input,
                    "expected": expected_sinhala,
                    "actual": actual_output,
                    "status": current_status,
                    "screenshot_rel": screenshot_rel,
                    "screenshot_name": screenshot_name,
                }
            )

        if args.keep_open and not args.headless:
            try:
                wb.save(args.output)
            except Exception:
                pass
            print("Keeping browser open. Press CTRL+C to stop.")
            try:
                while True:
                    page.wait_for_timeout(1000)
            except KeyboardInterrupt:
                try:
                    wb.save(args.output)
                except Exception:
                    pass
        browser.close()

    try:
        wb.save(args.output)
    except Exception as e:
        print(f"Error saving output file '{args.output}': {e}")
        return

    report_path = results_dir / "report.html"
    try:
        report_html = _build_report_html("SwiftTranslator Test Report", summary, report_rows)
        report_path.write_text(report_html, encoding="utf-8")
        print(f"HTML report saved to {report_path}")
    except Exception as e:
        print(f"Error writing HTML report '{report_path}': {e}")

    print(f"Test completed. Results saved to {args.output}")

if __name__ == "__main__":
    run_test()
